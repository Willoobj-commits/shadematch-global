#!/usr/bin/env python3
"""Build the Flutter seed catalogue from the source workbook.

The generated colour is deliberately a *universal-profile sample*. It is a
navigation aid, not a claim that the product has been colour measured.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


UNDERTONES: dict[str, dict[str, Any]] = {
    "N": {"name": "True Neutral", "vector": [0.0, 0.0], "lab": [0.0, 0.0]},
    "C": {"name": "Cool", "vector": [-1.0, 0.4], "lab": [3.0, -4.0]},
    "W": {"name": "Warm", "vector": [1.0, 0.0], "lab": [1.0, 5.0]},
    "G": {"name": "Golden", "vector": [1.1, -0.1], "lab": [0.5, 6.0]},
    "CP": {"name": "Cool Pink / Rosy", "vector": [-1.0, 1.0], "lab": [6.0, -5.0]},
    "CR": {"name": "Cool Red / Reddish", "vector": [-0.6, 1.2], "lab": [7.0, -1.0]},
    "CB": {"name": "Cool Blue", "vector": [-1.2, -0.2], "lab": [4.0, -7.0]},
    "WY": {"name": "Warm Yellow", "vector": [1.3, -0.2], "lab": [0.0, 8.0]},
    "WG": {"name": "Warm Golden", "vector": [1.2, -0.1], "lab": [1.0, 7.0]},
    "WP": {"name": "Warm Peach", "vector": [0.8, 0.8], "lab": [4.0, 4.0]},
    "PCH": {"name": "Peach", "vector": [0.5, 1.0], "lab": [5.0, 2.0]},
    "NP": {"name": "Neutral Peach", "vector": [0.2, 0.7], "lab": [3.0, 1.0]},
    "NG": {"name": "Neutral Golden", "vector": [0.4, -0.1], "lab": [0.0, 3.0]},
    "NR": {"name": "Neutral Red", "vector": [0.1, 0.9], "lab": [5.0, 0.0]},
    "WN": {"name": "Warm Neutral", "vector": [0.5, 0.0], "lab": [0.5, 2.5]},
    "CN": {"name": "Cool Neutral", "vector": [-0.5, 0.2], "lab": [1.5, -2.0]},
    "O": {"name": "Olive", "vector": [0.0, -1.0], "lab": [-3.0, 1.0]},
    "ON": {"name": "Neutral Olive", "vector": [0.0, -0.7], "lab": [-2.0, 0.5]},
    "OY": {"name": "Warm / Yellow Olive", "vector": [0.6, -0.8], "lab": [-2.0, 4.0]},
    "OG": {"name": "Golden Olive", "vector": [0.4, -1.0], "lab": [-3.0, 3.0]},
    "U": {"name": "Undertone Unmapped", "vector": None, "lab": [0.0, 0.0]},
}


def stable_id(prefix: str, *parts: Any) -> str:
    value = "|".join("" if part is None else str(part).strip().casefold() for part in parts)
    return f"{prefix}_{hashlib.sha1(value.encode('utf-8')).hexdigest()[:16]}"


def depth_family(depth: int) -> str:
    ranges = (
        (2, "Ultra Fair"), (5, "Fair"), (8, "Light"),
        (11, "Light-Medium"), (14, "Medium"), (17, "Tan"),
        (20, "Medium-Deep"), (24, "Deep"), (27, "Very Deep"),
        (30, "Deepest"),
    )
    return next(name for maximum, name in ranges if depth <= maximum)


def profile_lab(depth: int, undertone: str) -> tuple[float, float, float]:
    fraction = (depth - 1) / 29
    lightness = 94.0 - fraction * 72.0
    base_a = 7.0 + fraction * 7.0
    base_b = 13.0 + fraction * 9.0
    delta_a, delta_b = UNDERTONES.get(undertone, UNDERTONES["U"])["lab"]
    return round(lightness, 3), round(base_a + delta_a, 3), round(base_b + delta_b, 3)


def lab_to_hex(lab: tuple[float, float, float]) -> str:
    lightness, a_value, b_value = lab
    fy = (lightness + 16.0) / 116.0
    fx = a_value / 500.0 + fy
    fz = fy - b_value / 200.0

    def pivot_inverse(value: float) -> float:
        cube = value**3
        return cube if cube > 0.008856 else (value - 16.0 / 116.0) / 7.787

    x = 0.95047 * pivot_inverse(fx)
    y = 1.00000 * pivot_inverse(fy)
    z = 1.08883 * pivot_inverse(fz)
    linear = (
        x * 3.2404542 + y * -1.5371385 + z * -0.4985314,
        x * -0.9692660 + y * 1.8760108 + z * 0.0415560,
        x * 0.0556434 + y * -0.2040259 + z * 1.0572252,
    )

    def gamma(value: float) -> int:
        value = 12.92 * value if value <= 0.0031308 else 1.055 * value ** (1 / 2.4) - 0.055
        return max(0, min(255, round(value * 255)))

    red, green, blue = (gamma(value) for value in linear)
    return f"#{red:02X}{green:02X}{blue:02X}"


def clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def rows_as_dicts(sheet: Any) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    return [
        {str(header): clean(value) for header, value in zip(headers, row) if header}
        for row in rows
        if any(clean(value) is not None for value in row)
    ]


def build(source: Path, registry_js: Path | None = None) -> dict[str, Any]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    master = rows_as_dicts(workbook["MASTER_SHADES"])
    directory = rows_as_dicts(workbook["BRAND_REGISTRY"])
    source_register = rows_as_dicts(workbook["SOURCE_REGISTER"])
    source_lookup = {
        (
            str(row.get("Brand") or "").casefold(),
            str(row.get("Product") or "").casefold(),
            str(row.get("Official URL") or ""),
        ): row
        for row in source_register
    }
    source_by_url = {
        str(row.get("Official URL") or ""): row
        for row in source_register
        if row.get("Official URL")
    }
    shades: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    for row in master:
        brand = str(row["Brand"])
        product = str(row["Product Line"])
        market = str(row["Market Scope"])
        code = str(row["Shade Code"])
        name = clean(row.get("Shade Name"))
        depth = int(row["Universal Depth (1-30)"])
        undertone = str(row["Universal Undertone Code"] or "U").upper()
        if undertone not in UNDERTONES:
            raise ValueError(f"Unsupported undertone {undertone!r} for {brand} {code}")

        shade_id = stable_id("shd", brand, product, market, code, name)
        if shade_id in seen_ids:
            raise ValueError(f"Duplicate stable shade id: {brand} {product} {code}")
        seen_ids.add(shade_id)
        brand_id = stable_id("brd", brand)
        product_id = stable_id("prd", brand, product, market)
        lab = profile_lab(depth, undertone)
        display_hex = lab_to_hex(lab)
        source_url = str(row["Official Source URL"])
        source_metadata = source_lookup.get(
            (brand.casefold(), product.casefold(), source_url)
        ) or source_by_url.get(source_url, {})

        flags = ["profile_colour_estimate_only", "no_official_visual_asset"]
        if not row.get("Manufacturer Depth"):
            flags.append("manufacturer_depth_unavailable")
        if not row.get("Manufacturer Undertone"):
            flags.append("manufacturer_undertone_unavailable")
        if undertone == "U":
            flags.append("undertone_unmapped")
        if "current" not in str(row["Status"]).casefold():
            flags.append("not_current")

        visual_id = stable_id("vis", shade_id, "universal_profile_estimate")
        shades.append(
            {
                "id": shade_id,
                "brandId": brand_id,
                "brandName": brand,
                "origin": row["Country/Origin"],
                "marketScope": market,
                "productId": product_id,
                "productName": product,
                "productType": row["Product Type"],
                "shadeCode": code,
                "shadeName": name,
                "manufacturerDepth": row.get("Manufacturer Depth"),
                "manufacturerUndertone": row.get("Manufacturer Undertone"),
                "universalDepth": depth,
                "depthFamily": row.get("Universal Depth Family") or depth_family(depth),
                "undertoneCode": undertone,
                "undertoneName": row.get("Universal Undertone") or UNDERTONES[undertone]["name"],
                "universalProfile": row["Universal Profile"],
                "status": row["Status"],
                "normalizationBasis": row["Normalization Basis"],
                "sourceUrl": source_url,
                "sourceType": source_metadata.get("Source Type") or "Unclassified source",
                "sourceCoverageNote": source_metadata.get("Coverage Note"),
                "verifiedDate": str(row["Verified Date"]),
                "qualityFlags": flags,
                "visualReferences": [
                    {
                        "id": visual_id,
                        "kind": "universal_profile_estimate",
                        "label": "Universal profile colour sample",
                        "displayHex": display_hex,
                        "lab": {"l": lab[0], "a": lab[1], "b": lab[2]},
                        "colourSpace": "CIELAB_D65_estimate",
                        "measurementMethod": "derived_from_depth_and_undertone",
                        "verificationStatus": "system_generated",
                        "rightsStatus": "generated_in_app",
                        "matchEligible": False,
                        "isPrimary": True,
                        "sourceUrl": source_url,
                        "disclaimer": "Profile sample only; not an official or measured product swatch.",
                    }
                ],
            }
        )

    registry_metadata = {
        str(row["Brand"]).casefold(): row for row in directory if row.get("Brand")
    }
    registry_names = [str(row["Brand"]) for row in directory if row.get("Brand")]
    if registry_js:
        match = re.search(
            r"window\.GLOBAL_BRAND_REGISTRY=(\[.*?\]);",
            registry_js.read_text(encoding="utf-8"),
            flags=re.DOTALL,
        )
        if not match:
            raise ValueError(f"Could not find GLOBAL_BRAND_REGISTRY in {registry_js}")
        registry_names.extend(json.loads(match.group(1)))

    # Canonicalize case-only duplicates (the website registry contained both
    # MISSHA and Missha) while preferring the spelling used by populated data.
    populated_spelling = {shade["brandName"].casefold(): shade["brandName"] for shade in shades}
    canonical_names: dict[str, str] = {}
    aliases: list[dict[str, str]] = []
    for raw_name in registry_names:
        key = raw_name.casefold()
        preferred = populated_spelling.get(key, raw_name)
        if key in canonical_names and raw_name != canonical_names[key]:
            aliases.append({"alias": raw_name, "canonicalName": preferred})
        canonical_names[key] = preferred

    directory_items: list[dict[str, Any]] = []
    populated = {shade["brandName"] for shade in shades}
    for key, name in sorted(canonical_names.items(), key=lambda item: item[1].casefold()):
        row = registry_metadata.get(key, {})
        directory_items.append(
            {
                "id": stable_id("brd", name),
                "name": name,
                "origin": row.get("Country/Origin"),
                "primaryRegion": row.get("Primary Region"),
                "dataStatus": "verified" if name in populated else "queued",
                "sourceStatus": row.get("Shade Data Status") or "Directory / awaiting shade data",
            }
        )

    return {
        "schemaVersion": "2.0.0",
        "release": {
            "name": "Global v1 / Flutter visual-schema seed",
            "verifiedThrough": "2026-08-31",
            "shadeCount": len(shades),
            "verifiedBrandCount": len(populated),
            "directoryBrandCount": len(directory_items),
            "sourceFile": source.name,
        },
        "visualPolicy": {
            "fallbackKind": "universal_profile_estimate",
            "fallbackIsMatchEvidence": False,
            "requiredUserLabel": "Profile sample — not an official product swatch",
            "preferredEvidenceOrder": [
                "spectrophotometer_measurement",
                "calibrated_standardized_swatch",
                "official_digital_swatch",
                "official_product_image",
                "universal_profile_estimate",
            ],
        },
        "undertones": [
            {"code": code, "name": values["name"], "vector": values["vector"]}
            for code, values in UNDERTONES.items()
        ],
        "brandDirectory": directory_items,
        "brandAliases": aliases,
        "shades": shades,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--registry-js", type=Path)
    args = parser.parse_args()
    catalogue = build(args.source, args.registry_js)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(catalogue, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(json.dumps(catalogue["release"], indent=2))


if __name__ == "__main__":
    main()
